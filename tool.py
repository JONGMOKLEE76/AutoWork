import pandas as pd
import numpy as np
import os
import re
import datetime
from dateutil.relativedelta import relativedelta
from tkinter import *
from tkinter import filedialog
from mapping import *
import requests
import openpyxl
import pickle
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tqdm import tqdm
from pandas.api.types import CategoricalDtype
import matplotlib.pyplot as plt
from google.cloud import translate_v2 as translate
import pyautogui
import sys

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = r'C:\Users\paul76.lee\folkloric-pier-362414-cf467b42237d.json'
odm_model = re.compile(r'\b1\d[A-Z][A-Z]?\d\d[A-Z]-[A-Z]\.[A,C]\w{4,5}\b') # LG PC ODM 모델명의 정규표현식
odm_sn = re.compile('^\d{3}[Q, P][C, G][A-Z]{2}\d{6}$|^\d{8}[Q, P][C, G][A-Z]{2}\d{6}$') # Quanta / Pegatron 모델의 S/N의 정규표현식

def check_model_name(name):
    m = odm_model.match(name)
    if m:
        return True
    else:
        return m

def getfirstdate_year(date):
    firstdayofweek = date - datetime.timedelta(days = date.isocalendar()[2] - 1)
    return '{}'.format(firstdayofweek.year)

def getfirstdate_month(date):
    firstdayofweek = date - datetime.timedelta(days = date.isocalendar()[2] - 1)
    return '{0:02d}'.format(firstdayofweek.month)

def getfirstdate_day(date):
    firstdayofweek = date - datetime.timedelta(days = date.isocalendar()[2] - 1)
    return '{0:02d}'.format(firstdayofweek.day)

def get_month_from_date(date):
    month_list = []
    for i in range(7):
        month_list.append((date - datetime.timedelta(days=date.isocalendar()[2]-1-i)).month)
    month_list = pd.Series(month_list)
    max_value = month_list.value_counts().max()
    for i in month_list.value_counts().index:
        if month_list.value_counts().loc[i] == max_value:
            return i

def get_weekname(date):
    firstdayofweek = date - datetime.timedelta(days = date.isocalendar()[2] - 1)
    return f"{firstdayofweek.year}-{firstdayofweek.month:02d}-{firstdayofweek.day:02d}(W{date.isocalendar()[1]:02d})"

# week이름으로 표시된 데이터프레임의 컬럼을 월별 합계로 변환하는 함수
def category_sum(df):
    p = re.compile('-\d\d-')
    temp_df = pd.DataFrame(index=df.index)
    for col_name in df.columns:
        if type(col_name) == str:
            if p.search(col_name) != None:
                date = datetime.date.fromisoformat(col_name[:10])
                value = '{0}-{1:02d}'.format(date.isocalendar().year,get_month_from_date(date))
                if value in temp_df.columns:
                    temp_df[value] += df[col_name]
                else:
                    temp_df[value] = df[col_name]
    return temp_df

def get_weeklist(listobj):
    weeklist = [i.replace('\n','') for i in listobj if type(i)==str]
    weeklist = [i for i in weeklist if re.compile('\d\d\(W\d').search(i)!=None]
    return weeklist

def trim_column_name(df):
    for name in df.columns:
        if type(name) == str:
            if re.compile('(W\d\d)').search(name) != None:
                df.rename(columns={name:re.sub('\r?\n', '', name)}, inplace=True)

# 특정 list에서 원하는 패턴의 문자열이 있는 것만 추출하여 리스트로 반환하는 함수
def get_pattern_from_list(any_list, pattern):
    desired_list = []
    for item in any_list:
        if type(item) == str:
            if pattern.search(item) != None:
                desired_list.append(item)
    return desired_list

def get_PC_ODM_shipment_plan(filename, vendor_name, sheet):
    df = pd.read_excel(filename, sheet_name=sheet, skiprows = 5, na_values=['', 'nan'], keep_default_na=False) # sheet name을 번호로 설정하면 원하는 Sheet를 가져올 수 있음. 순서는 '0' 부터
    df = df[df['From Site'].notnull()]
    df = df[df['From Site'].str.contains(vendor_name)] # ODM vendor명에 따라 필터링
    idx_num = df.columns.get_loc('Category')
    col1 = df.columns.tolist()[3:idx_num+1] # 필요한 컬럼만 리스트로 정리

    col2 = [] # 컬럼명 중에서 각 주차에 해당하는 컬럼명만 추출
    for item in df.columns.tolist()[idx_num+1:]:
        if type(item) == str:
            if re.compile('(W\d\d)').search(item) != None:
                df.rename(columns = {item:re.sub('\r?\n', '', item)}, inplace = True)
                item = re.sub('\r?\n', '', item)
                col2.append(item)
    df[col2] = df[col2].fillna(0)
    df = df[col1+col2] # 원하는 컬럼만 선택하여 데이타프레임 재구성
    df['Model'] = df['Model.Suffix'].apply(lambda x: x.split('-')[0])
    df['Series'] = df['Model'].replace(srt_model)
    df['Sum'] = df[col2].sum(axis=1)
    df = df[df['Sum'] != 0] # 물동이 '0' 인 행은 제외
    df.drop(columns='Sum', inplace=True)
    df.reset_index(drop=True, inplace=True) # index 재설정
    for key, val in df.dtypes.items(): # df의 dtype이 float 인 것을 int로 모두 변경
        if val == 'float64':
            p = re.compile('-\d\d-')
            if p.search(str(key)) != None:
                df[key] = df[key].astype('int')
    return df

# PDR 정보를 데이타프레임으로 생성하는 함수
def get_pdr():
    os.chdir('D:/Python result/Quanta PDRDB/')
    filelist = os.listdir() # 다운받은 폴더의 내용을 리스트로 만듬

    pdr = pd.DataFrame()

    for file in filelist:
#         temp = pd.read_csv(file, encoding='cp949')
        temp = pd.read_csv(file)
        pdr = pd.concat([pdr, temp])
    pdr.insert(0, 'Model', pdr['Model.Suffix'].apply(lambda x:x.split('-')[0]))
    pdr.insert(0, 'Series', pdr['Model.Suffix'].apply(lambda x:x.split('-')[0]).replace(srt_model))
    pdr.reset_index(inplace=True, drop=True)
    os.chdir('C:/Users/paul76.lee/')
    return pdr

# SP를 필요한 정보를 담은 컬럼만 구성하여 DataFrame으로 만드는 함수
def trim_sp(df):
    week = get_weeklist(df)
    df['Model'] = df['Model.Suffix'].apply(lambda x:x.split('-')[0])
    df['Series'] = df['Model'].replace(srt_model)
    basic_col = ['Model.Suffix', 'Series', 'Model', 'Region', 'Subsidiary', 'To Site', 'From Site', 'Transport Mode']
    df = df[basic_col+week]
    return df

#2개의 DataFrame 의 공통 컬럼의 값의 차이(gap)을 계산하여 DataFrame으로 Return 하는 함수
def cal_gap(df1, df2):
    df1 = df1.set_index(['Model.Suffix', 'Series', 'Model', 'Region', 'Subsidiary', 'To Site','From Site', 'Transport Mode'])
    df2 = df2.set_index(['Model.Suffix', 'Series', 'Model', 'Region', 'Subsidiary', 'To Site','From Site', 'Transport Mode'])
    u1 = set(df1.columns) | set(df2.columns)
    u2 = pd.concat([pd.DataFrame(index = df1.index), pd.DataFrame(index = df2.index)], axis = 1).index
    df = pd.DataFrame(index = u2)

    for col in u1:
        c = []
        for ind in u2:
            try:
                a = df1.loc[ind, col]
            except:
                a = 0
            try:
                b = df2.loc[ind, col]
            except:
                b = 0
            c.append(a-b)
        df[col] = c
    a = df.columns.tolist()
    a.sort()
    return df[a]
    

# 2개의 Dataframe의 특정 컬럼의 값에 대한 SP값의 차이를 구하는 함수
def cal_gap2(df1, df2, col_name):
    ttl_row = set(df1[col_name].unique()) | set(df2[col_name].unique())
    df1_wk = get_weeklist(df1)
    df2_wk = get_weeklist(df2)
    ttl_col = set(df1_wk) | set(df2_wk)

    gap_df = pd.DataFrame()

    for item in ttl_row:
        filt1 = (df1[col_name] == item)
        filt2 = (df2[col_name] == item)
        for wk_nm in ttl_col:
            try:
                a = df1.loc[filt1, wk_nm].sum() 
            except:
                a = 0
            try:
                b = df2.loc[filt2, wk_nm].sum()
            except:
                b = 0
            gap_df.loc[item,wk_nm] = a - b
    gap_df.columns = gap_df.columns.sort_values()
    gap_df = category_sum(gap_df)
    col_list = gap_df.columns.to_list()
    for col in get_weeklist(gap_df):
        if col in col_list:
            col_list.remove(col)
    gap_df = gap_df[col_list]
    return gap_df

# 데이타프레임의 특정 컬럼값에 대해 grouping 한 결과를  월별로 합산한 결과를 계산하는 함수
def monthly_sum(df, col_nm):
    new_df = pd.DataFrame(index=df.index)
    df = category_sum(df.groupby(by=col_nm).sum(numeric_only=True))
    wk = get_pattern_from_list(df.columns, re.compile('-\d\d-'))
    col_list = sorted(list(set(df.columns) - set(wk)))
    df = df[col_list]
    return df

def get_filepath(word):
    os.chdir('D:/Downloads/') # Shipment plan 다운받은 폴더로 이동
    filelist = os.listdir() # 다운받은 폴더의 내용을 리스트로 만듬
    match = [i for i in filelist if word in i] # 상기의 리스트에서 특정 단어가 들어간 파일명만 추출하여 리스트화
    path = 'D:/Downloads/' + match[0] # DataFrame으로 만들 엑셀 파일의 경로명을 설정
    return path

def make_dict_from_df(dataframe):
    return dict(zip(dataframe.iloc[:,0], dataframe.iloc[:,1]))

def get_shipment_result_from_GLOP(row_list, col_list):
    path = get_filename()
    df = pd.read_excel(path).iloc[:, :48]
    df = df[df['Shipping'] == 'Y']
    df = df.drop(['Biz Type', 'Rev. ORG', 'PO Status', 'Receipt', 'Cancel', 'Shipping', 'Demand', 'saNo', 'tcNo', 'ccNo', 'OQC Report', 'OQC Result', 'Container   Insp&Report'], axis = 1)
    df = df.convert_dtypes()
    df['Ship Date'] = pd.to_datetime(df['Ship Date'])
    df['Issued Date'] = pd.to_datetime(df['Issued Date'])
    df['PO Week'] = df['Issued Date'].dt.isocalendar().week.apply(lambda x:'{0:02d}'.format(x))
    df['PO Month'] = df['Issued Date'].apply(get_month_from_date)
    df['PO Year'] = df['Issued Date'].dt.isocalendar().year

    df['RSD'] = pd.to_datetime(df['RSD'])
    df['RSD Week'] = df['RSD'].dt.isocalendar().week.apply(lambda x:'{0:02d}'.format(x))
    df['RSD Month'] = df['RSD'].apply(get_month_from_date)
    df['RSD Year'] = df['RSD'].dt.isocalendar().year
    df['RSD Week Year'] = df['RSD'].apply(getfirstdate_year)
    df['RSD Week Month'] = df['RSD'].apply(getfirstdate_month)
    df['RSD Week Day'] = df['RSD'].apply(getfirstdate_day)

    df['RSD Week Name'] = df['RSD Week Year'].astype(str) + '-' + df['RSD Week Month'].astype(str) + '-' + df['RSD Week Day'].astype(str) + '(W' + df['RSD Week'].astype(str) +')'

    df['Expected Ship Date'] = pd.to_datetime(df['Expected Ship Date'])
    df['Ship Week'] = df['Ship Date'].dt.isocalendar().week.apply(lambda x:'{0:02d}'.format(x))
    df['Ship Month'] = df['Ship Date'].apply(get_month_from_date)
    df['Year'] = df['Ship Date'].dt.isocalendar().year
    df.rename(columns={'Model':'Model.Suffix'}, inplace=True)
    df['Model'] = df['Model.Suffix'].apply(lambda x: x.split('-')[0])
    df['Series'] = df['Model'].replace(srt_model)
    df['Country'] = df['Ship To'].replace(site_map)
    df['Region'] = df['Country'].replace(country_map)
    df['Ship Week month'] = df['Ship Date'].apply(getfirstdate_month)
    df['Ship Week day'] = df['Ship Date'].apply(getfirstdate_day)
    df['Ship Week year'] = df['Ship Date'].apply(getfirstdate_year)
    df['Ship Week year'] = df['Ship Week year'].astype(str)
    df['Ship Week month'] = df['Ship Week month'].astype(str)
    df['Ship Week day'] = df['Ship Week day'].astype(str)
    df['Ship Week'] = df['Ship Week'].astype(str)
    df['Week Name'] = df['Ship Week year'] + '-' + df['Ship Week month'] + '-' + df['Ship Week day'] + '(W' + df['Ship Week'] + ')'
    df_pv = pd.pivot_table(df, index = row_list, columns = col_list, values = ['Ship'], aggfunc = 'sum').fillna(0)
    return df_pv

def get_qt_production_plan():
    path = get_filename()
    pd_df = pd.read_excel(path)
    cond = (pd_df.iloc[:, 0] == 'MODEL')
    i_num = pd_df.loc[cond].index[0]
    pd_df = pd.read_excel(path, skiprows=i_num+1)
    pd_df = pd_df.drop('MODEL', axis = 1)
    pd_df = pd_df[pd_df['LGE P/N'].notnull()]
    pd_df.rename(columns={'LGE P/N':'Model.Suffix'}, inplace=True)
    pd_df = pd_df.loc[:, :'Balance']
    pd_df = pd_df[~pd_df['Model.Suffix'].str.contains('total')]
    pd_df = pd_df.drop(['MB', 'OS', 'LCD', 'SSD', 'HDD','Bat'], axis = 1)
    pd_df.dropna(axis=1, how='all', inplace=True)
    pd_df.fillna(0, axis=1, inplace=True)
    pd_df['Model'] = pd_df['Model.Suffix'].apply(lambda x:x.split('-')[0])
    pd_df['Series'] = pd_df['Model'].replace(srt_model)
    pd_df.astype({'Input Plan':'int'})
    pd_df = pd_df.reset_index(drop=True)
    return pd_df

def get_google_translate(text, lan='en'):
    translator = translate.Client()
    return translator.translate(text, target_language=lan)['translatedText']

def get_papago_translate(text, lan):
    ''' 언어코드 참조(파파고)
        ko  한국어
        en  영어
        ja  일본어
        zh-CN   중국어 간체
        zh-TW   중국어 번체
        vi  베트남어
        id  인도네시아어
        th  태국어
        de  독일어
        ru  러시아어
        es  스페인어
        it  이탈리아어
        fr  프랑스어 '''
    client_id = "DY2EGB7Wjy586X0xyYkR"
    client_secret = "mjQv54VXNG"
    
    data = {'text' : text,
            'source' : 'ko',
            'target': lan}

    url = "https://openapi.naver.com/v1/papago/n2mt"

    header = {"X-Naver-Client-Id":client_id,
              "X-Naver-Client-Secret":client_secret}

    response = requests.post(url, headers=header, data= data)
    rescode = response.status_code

    if(rescode==200):
        t_data = response.json()
        return t_data['message']['result']['translatedText']
    else:
        print("Error Code:" , rescode)

def get_filename():
    root = Tk()
    filename = filedialog.askopenfilename()
    root.mainloop()
    return filename

def get_difference_table(dataframe1, dataframe2, pattern):
    df1 = dataframe1.copy()
    df2 = dataframe2.copy()
    df1_col = []
    df1_wk_list = get_pattern_from_list(df1.columns, re.compile(pattern))
    for col_name in df1.columns:
        if col_name not in df1_wk_list:
            df1_col.append(col_name)
    df2_col = []
    df2_wk_list = get_pattern_from_list(df2.columns, re.compile(pattern))
    for col_name in df2.columns:
        if col_name not in df2_wk_list:
            df2_col.append(col_name)

    if len(df1_col) != 0:
        df1.set_index(df1_col, inplace=True)
    if len(df2_col) != 0:
        df2.set_index(df2_col, inplace=True)

    joined_col = list(set(df1.columns) | set(df2.columns))
    joined_col.sort()
    
    m_index = list(set(df1.index)|set(df2.index))
    if type(m_index[0]) == tuple:
        m_index = pd.MultiIndex.from_tuples(m_index)
    df_diff_table = pd.DataFrame(data=0, index=m_index, columns=joined_col).sort_index()
    df_diff_table.index.names = df1.index.names

    for i in df_diff_table.index:
        for j in df_diff_table.columns:
            try:
                df1_value = df1.loc[i,j]
            except:
                df1_value = 0
            try:
                df2_value = df2.loc[i,j]
            except:
                df2_value = 0
            df_diff_table.loc[i,j] = df1_value - df2_value
#     if len(df1_col) != 0:
#         df_diff_table.index.names = df1_col
    return df_diff_table

# 엑셀의 특정 PDR sheet 1개를 데이타프레임으로 변환
def convert_df_from_PDR(path, sheetnum):
    df = pd.read_excel(path, sheet_name = sheetnum, header = None)
    df1 = df[[0, 1]].dropna().rename(columns = {0:'item', 1:'value'})
    df2 = df[[2, 3]].dropna().rename(columns = {2:'item', 3:'value'})
    df3 = df[[4, 5]].dropna().rename(columns = {4:'item', 5:'value'})
    df_result = pd.concat([df1, df2, df3], axis = 0).reset_index(drop=True)
    df_result = df_result.T
    df_result = df_result.rename(columns=df_result.iloc[0])
    df_result.drop('item', inplace=True)
    df_result.set_index('Model.Suffix', inplace=True)
    return df_result

# 여러 SKU의 PDR이 담겨있는 엑셀 파일에서 전체 SKU의 PDR을 하나의 데이타프레임으로 만듬
def make_all_PDR_data_from_certain_model(filename):
    df = pd.DataFrame()
    wb = openpyxl.load_workbook(filename)
    for i in range(2, len(wb.sheetnames)): # 엑셀 파일의 2번째 sheet부터 PDR 이 있으므로
        temp = convert_df_from_PDR(filename, i)
        df = pd.concat([df, temp])
    wb.close()
    return df

def google_translate(text): 
    browser = webdriver.Chrome()
    url = 'https://translate.google.co.kr/?hl=ko&sl=auto&tl=en&text='
    browser.get(url)
    elem = browser.find_element_by_tag_name('textarea')
    elem.send_keys(text)
    elem = WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.CLASS_NAME, 'J0lOec')))
    result = elem.text
    browser.quit()
    return result

def get_weeklist_for_certain_month(year, month):
    weeklist = []
    first_date = datetime.date(year, month, 1)
    for i in range(5):
        each_date = first_date + datetime.timedelta(days=i*7)
        if get_month_from_date(each_date) == month:
            weeklist.append(get_weekname(each_date))
    return weeklist

def get_weekname_from(weekname, num):
    return get_weekname(datetime.date.fromisoformat(weekname[:10]) + datetime.timedelta(days=7)*num)


def get_open_po_at_certainweek(vendor, weekname):
    with open(f'D:/Data/{vendor} PO DB.bin', 'rb') as f:
        PO_df = pickle.load(f)
    with open(f'D:/Data/{vendor} shipment result DB.bin', 'rb') as f:
        SR_df = pickle.load(f)

    open_po_info = {} # Open PO 정보를 dictionary 형태로 저장
    grouped = SR_df.groupby('PO No.')
    for i in PO_df.index:
        try:
            cond = grouped.get_group(i)['Week Name'] <= weekname
            shipped_qty = grouped.get_group(i).loc[cond, 'Ship'].sum()
        except:
            shipped_qty = 0
        open_po_qty = PO_df.loc[i, 'PO'] - shipped_qty
        if open_po_qty > 0:
            open_po_info[i] = open_po_qty

    open_po_df = PO_df.loc[open_po_info.keys()]
    open_po_df['Open_Qty'] = open_po_info.values()
    open_po_df = open_po_df[open_po_df['Cancel'] == 'N']
    return open_po_df

def get_open_po(vendor):
    with open(f'D:/Data/{vendor} PO DB.bin', 'rb') as f:
        PO_df = pickle.load(f)
    with open(f'D:/Data/{vendor} shipment result DB.bin', 'rb') as f:
        SR_df = pickle.load(f)

    open_po_info = {} # Open PO 정보를 dictionary 형태로 저장
    grouped = SR_df.groupby('PO No.')
    for i in PO_df.index:
        try:
            shipped_qty = grouped.get_group(i)['Ship'].sum()
        except:
            shipped_qty = 0
        open_po_qty = PO_df.loc[i, 'PO'] - shipped_qty
        if open_po_qty > 0:
            open_po_info[i] = open_po_qty

    open_po_df = PO_df.loc[open_po_info.keys()]
    open_po_df['Open_Qty'] = open_po_info.values()
    open_po_df = open_po_df[open_po_df['Cancel'] == 'N']
    return open_po_df

def get_lastmonth(weekname):
    th_mon = get_month_from_date(datetime.date.fromisoformat(weekname[:10]))
    th_year = datetime.date.fromisoformat(weekname[:10]).isocalendar()[0]
    if th_mon == 1:
        month = 12
        year = th_year - 1
    else:
        month = th_mon - 1
        year = th_year
    return (year, month)

def adjust_shipment_plan(vendor, thisweek, ver='Final'):
    lastweek = get_weekname_from(thisweek, -1)
    lastmonth = get_lastmonth(thisweek)

    # 이번주 기준 전달까지의 이전주차 목록만들기
    weeklist1 = get_weeklist_for_certain_month(lastmonth[0], lastmonth[1])
    weeklist2 = get_weeklist_for_certain_month(datetime.date.fromisoformat(thisweek[:10]).isocalendar()[0] , get_month_from_date(datetime.date.fromisoformat(thisweek[:10])))
    if weeklist2.index(thisweek) == 0:
        previous_weeklist = weeklist1
    else:
        previous_weeklist = weeklist1 + weeklist2[:weeklist2.index(thisweek)]

    if vendor == 'Wanlida':
        with open(f'D:/Data/{vendor} shipment result DB.bin', 'rb') as f:
            sr = pickle.load(f)
        sr = sr[sr['Week Name'].isin(previous_weeklist)]
        sr1 = sr.groupby(['Mapping Model.Suffix', 'Ship To', 'Week Name'])['Ship'].sum()
        sr2 = pd.DataFrame(data=0, columns=previous_weeklist, index=sr.groupby(['Mapping Model.Suffix', 'Ship To']).sum().index)
        for x, y, z in sr1.index:
            sr2.loc[(x, y), z] = sr1[(x,y,z)]
        sr2.index.names = ['Mapping Model.Suffix', 'To Site']

        sp = pd.read_excel(get_filename(), sheet_name='Shipment Plan', keep_default_na=False, na_values=['', 'nan']) # GSCP로부터 금주 SP를 데이타프레임화
        trim_column_name(sp)
        sp = sp[sp['From Site'].str.contains(vendor.upper())]
        sp = sp[sp['Category'].str.contains(ver)]
        sp = sp[['Mapping Model.Suffix', 'To Site']+get_pattern_from_list(sp.columns, re.compile('-\d\d-'))].reset_index(drop=True)
        sp = sp.convert_dtypes() # 적절한 데이타타입으로 변경
        sp['To Site'] = sp['To Site'].replace(site_name_adjust_map)
        sp1 = sp.groupby(['Mapping Model.Suffix', 'To Site']).sum()
        sp2 = sp1.loc[:, :thisweek]
        sp2['Sum'] = sp2.sum(axis=1)
        
        c1 = (sr['Cancel'] == 'N')
        c2 = (sr['RSD Week Name'] <= thisweek)
        sr3 = sr[c1 & c2].groupby(['Mapping Model.Suffix', 'Ship To'])['Ship'].sum()
        if sr3.shape[0] != 0:
            for i in sr3.index:
                sp2.loc[i, 'Shipped'] = sr3[i]
        else:
            sp2['Shipped'] = 0

        sp2 = sp2.fillna(0)
        sp2['Real SP'] = sp2['Sum'] - sp2['Shipped']
        sp2 = sp2['Real SP']
        sp2.name = thisweek
        sp2 = sp2.astype('int')

        if sr[c1&~c2].shape[0] == 0:
            print(f'{thisweek} No Pre-Ship in future orders.')
        else:
            print(f'{thisweek} Pre-Ship exits in future orders')
            prship = sr[c1&~c2]
            print(prship.pivot_table('Ship', index=['Mapping Model.Suffix', 'Ship To'], columns='RSD Week Name'))
            prship = prship.groupby(['Mapping Model.Suffix', 'Ship To', 'RSD Week Name'])['Ship'].sum().reset_index()

            print(f'이전 주 {lastweek} GSCP자료를 한번 더 열어주세요.')
            sp_lt = pd.read_excel(get_filename(), sheet_name='Shipment Plan', keep_default_na=False, na_values=['', 'nan'])
            trim_column_name(sp_lt)
            sp_lt = sp_lt[sp_lt['From Site'].str.contains(vendor.upper())]
            sp_lt = sp_lt[sp_lt['Category'].str.contains(ver)]
            sp_lt = sp_lt[['Mapping Model.Suffix', 'To Site']+get_pattern_from_list(sp_lt.columns, re.compile('-\d\d-'))].reset_index(drop=True)
            sp_lt = sp_lt.convert_dtypes() # 적절한 데이타타입으로 변경
            sp_lt['To Site'] = sp_lt['To Site'].replace(site_name_adjust_map)
            sp_lt = sp_lt.groupby(['Mapping Model.Suffix', 'To Site']).sum()
            sp_lt = sp_lt.convert_dtypes()

            for i in prship.iterrows():
                md = i[1]['Mapping Model.Suffix']
                st = i[1]['Ship To']
                wk = i[1]['RSD Week Name']
                
                lt_ship_qty = sp_lt.loc[(md, st), wk]
                th_ship_qty = sp1.loc[(md, st), wk]
                if lt_ship_qty - th_ship_qty == i[1]['Ship']: # 선행선적한 SP가 지난 주 기준의 해당주차의 SP값에서 이번주 기준의 SP값을 뺐을 때 이 값이 선행선적한 값과 같다면, 이번주 나온 SP가 선행선적 수량을 차감해서 제대로 나온 것이고
                    sp2.loc[(md, st)] = sp2.loc[(md, st)] - i[1]['Ship'] # 이번주 Real SP값 계산할 때, 선행 선적되어 반영된 SP값을 차감하여 보정하여 재계산해야 함
                    continue
                sp1.loc[(md, st), wk] = sp1.loc[(md, st), wk] - i[1]['Ship']

        sp3 = sp1.loc[:, get_weekname_from(thisweek, 1):]
        sp4 = pd.concat([sr2, sp2, sp3], axis=1).fillna(0)
        sp4 = sp4.reset_index()
        sp4.insert(1, 'Country', sp4['To Site'].replace(site_map))
        sp4.insert(0, 'Region', sp4['Country'].replace(country_map))
        sp4 = sp4.convert_dtypes()
        return sp4

    # ODM업체가 Wanlida가 아닐때
    else:
        # 해당 업체 선적실적 불러와서 이전주차들의 SP 취합
        with open(f'D:/Data/{vendor} shipment result DB.bin', 'rb') as f:
            sr = pickle.load(f)
        sr1 = pd.pivot_table(sr, index = ['Mapping Model.Suffix', 'Ship To'], columns = 'Week Name', values = 'Ship',
                        aggfunc=sum)[previous_weeklist].dropna(how='all').fillna(0).convert_dtypes()
        sr1.index.names = ['Mapping Model.Suffix', 'To Site']

        # 이번주의 SP 구하기
        with open(f'D:/Data/{vendor} forecast.bin', 'rb') as f:
            sp = pickle.load(f)
        sp = sp[sp['Ref_week'] == thisweek] # 이번 주 기준의 Fcst 가져오기
        sp['To Site'] = sp['To Site'].replace(site_name_adjust_map)
        sp1 = sp.groupby(['Mapping Model.Suffix', 'To Site']).sum()
        sp1 = sp1.loc[:, lastweek:]

        sp2 = sp1[[lastweek, thisweek]]
        sp2['Sum'] = sp2[lastweek] + sp2[thisweek]
        c1 = sr['Week Name'] == lastweek
        c2 = sr['Cancel'] == 'N'
        c3 = sr['RSD Week Name'] <= thisweek
        sr2 = sr[c1 & c2 & c3].groupby(['Mapping Model.Suffix', 'Ship To'])['Ship'].sum()
        for i in sr2.index:
            sp2.loc[i, 'Shipped'] = sr2[i]
        sp2 = sp2.fillna(0)
        sp2['Real SP'] = sp2['Sum'] - sp2['Shipped']
        sp2 = sp2['Real SP']
        sp2.name = thisweek
        sp2 = sp2.astype('int')

        if sr[c1&c2][sr[c1&c2]['RSD Week Name'] > thisweek].shape[0] == 0:
            print('No Pre-Ship in future orders.')
        else:
            print('Pre-Ship exits in future orders')
            prship = sr[c1&c2][sr[c1&c2]['RSD Week Name'] > thisweek]
            print(prship.pivot_table('Ship', index=['Mapping Model.Suffix', 'Ship To'], columns='RSD Week Name'))
            prship = prship.reset_index()
            for i in prship.iterrows():
                md = i[1]['Mapping Model.Suffix']
                st = i[1]['Ship To']
                wk = i[1]['RSD Week Name']
                with open(f'D:/Data/{vendor} forecast.bin', 'rb') as f:
                    last_sp = pickle.load(f)
                last_sp = last_sp[last_sp['Ref_week'] == lastweek] 
                last_sp = last_sp.groupby(['Mapping Model.Suffix', 'To Site']).sum()
                last_sp = last_sp.convert_dtypes()
                lt_ship_qty = last_sp.loc[(md, st), wk]
                th_ship_qty = sp1.loc[(md, st), wk]
                if lt_ship_qty - th_ship_qty == i[1]['Ship']:
                    sp2.loc[(md, st)] = sp2.loc[(md, st)] - i[1]['Ship'] # 이번주 Real SP값 계산할 때, 선행 선적되어 반영된 SP값을 차감하여 보정하여 재계산해야 함
                    continue
                sp1.loc[(md, st), wk] = sp1.loc[(md, st), wk] - i[1]['Ship']

        sp3 = sp1.iloc[:, 2:]
        sp4 = pd.concat([sr1, sp2, sp3], axis=1).fillna(0)
        sp4 = sp4.convert_dtypes()
        sp4 = sp4.reset_index()
        sp4.insert(1, 'Country', sp4['To Site'].replace(site_map))
        sp4.insert(1, 'Region', sp4['Country'].replace(country_map))
        sp4.insert(0, 'Series', sp4['Mapping Model.Suffix'].apply(lambda x:x.split('-')[0]).replace(srt_model))
        return sp4

def get_pc_shipment_plan_from_GSCP(thisweek, ver, vendor):
    print(f'이번주 {thisweek} 의 GSCP SP {ver} 를 가져옵니다.')
    lastweek = get_weekname_from(thisweek, -1)
    lastmonth = get_lastmonth(thisweek)

    weeklist1 = get_weeklist_for_certain_month(lastmonth[0], lastmonth[1])
    weeklist2 = get_weeklist_for_certain_month(datetime.date.fromisoformat(thisweek[:10]).isocalendar()[0] , get_month_from_date(datetime.date.fromisoformat(thisweek[:10])))
    if weeklist2.index(thisweek) == 0:
        previous_weeklist = weeklist1
    else:
        previous_weeklist = weeklist1 + weeklist2[:weeklist2.index(thisweek)]

    with open(f'D:/Data/{vendor} shipment result DB.bin', 'rb') as f:
        sr = pickle.load(f)
    sr = sr[sr['Week Name'].isin(previous_weeklist)]
    sr1 = sr.groupby(['Mapping Model.Suffix', 'Ship To', 'Week Name'])['Ship'].sum()
    sr2 = pd.DataFrame(data=0, columns=previous_weeklist, index=sr.groupby(['Mapping Model.Suffix', 'Ship To']).sum().index)
    for x, y, z in sr1.index:
        sr2.loc[(x, y), z] = sr1[(x,y,z)]
    sr2.index.names = ['Mapping Model.Suffix', 'To Site']

    sp = pd.read_excel(get_filename(), sheet_name='Shipment Plan', keep_default_na=False, na_values=['', 'nan']) # GSCP로부터 금주 SP를 데이타프레임화
    trim_column_name(sp)
    sp = sp[sp['From Site'].str.contains(vendor.upper())]
    sp = sp[sp['Category'].str.contains(ver)]
    sp = sp[['Mapping Model.Suffix', 'To Site']+get_pattern_from_list(sp.columns, re.compile('-\d\d-'))].reset_index(drop=True)
    sp = sp.convert_dtypes() # 적절한 데이타타입으로 변경
    sp['To Site'] = sp['To Site'].replace(site_name_adjust_map)
    sp1 = sp.groupby(['Mapping Model.Suffix', 'To Site']).sum()
    sp2 = sp1.loc[:, :thisweek]
    sp2['Sum'] = sp2.sum(axis=1)

    c1 = (sr['Cancel'] == 'N')
    c2 = (sr['RSD Week Name'] <= thisweek)
    sr3 = sr[c1 & c2].groupby(['Mapping Model.Suffix', 'Ship To'])['Ship'].sum()
    if sr3.shape[0] != 0:
        for i in sr3.index:
            sp2.loc[i, 'Shipped'] = sr3[i]
    else:
        sp2['Shipped'] = 0

    sp2 = sp2.fillna(0)
    sp2['Real SP'] = sp2['Sum'] - sp2['Shipped']
    sp2 = sp2['Real SP']
    sp2.name = thisweek

    if sr[c1&~c2].shape[0] == 0:
        print(f'{thisweek} No Pre-Ship in future orders.')
    else:
        print(f'{thisweek} Pre-Ship exits in future orders')
        prship = sr[c1&~c2]
        print(prship.pivot_table('Ship', index=['Mapping Model.Suffix', 'Ship To'], columns='RSD Week Name'))
        prship = prship.groupby(['Mapping Model.Suffix', 'Ship To', 'RSD Week Name'])['Ship'].sum().reset_index()

        print(f'이전 주{lastweek} GSCP자료를 한번 더 열어주세요.')
        sp_lt = pd.read_excel(get_filename(), sheet_name='Shipment Plan', keep_default_na=False, na_values=['', 'nan'])
        trim_column_name(sp_lt)
        sp_lt = sp_lt[sp_lt['From Site'].str.contains(vendor.upper())]
        sp_lt = sp_lt[sp_lt['Category'].str.contains('Final')]
        sp_lt = sp_lt[['Mapping Model.Suffix', 'To Site']+get_pattern_from_list(sp_lt.columns, re.compile('-\d\d-'))].reset_index(drop=True)
        sp_lt = sp_lt.convert_dtypes() # 적절한 데이타타입으로 변경
        sp_lt['To Site'] = sp_lt['To Site'].replace(site_name_adjust_map)
        sp_lt = sp_lt.groupby(['Mapping Model.Suffix', 'To Site']).sum()
        sp_lt = sp_lt.convert_dtypes()

        for i in prship.iterrows():
            md = i[1]['Mapping Model.Suffix']
            st = i[1]['Ship To']
            wk = i[1]['RSD Week Name']

            lt_ship_qty = sp_lt.loc[(md, st), wk]
            th_ship_qty = sp1.loc[(md, st), wk]
            if lt_ship_qty - th_ship_qty == i[1]['Ship']: # 선행선적한 SP가 지난 주 기준의 해당주차의 SP값에서 이번주 기준의 SP값을 뺐을 때 이 값이 선행선적한 값과 같다면, 이번주 나온 SP가 선행선적 수량을 차감해서 제대로 나온 것이고
                sp2.loc[(md, st)] = sp2.loc[(md, st)] - i[1]['Ship'] # 이번주 Real SP값 계산할 때, 선행 선적되어 반영된 SP값을 차감하여 보정하여 재계산해야 함
                continue
            sp1.loc[(md, st), wk] = sp1.loc[(md, st), wk] - i[1]['Ship'] # 위의 값이 다르다면, 실제 선행 선적은 했는데, 해당 수량을 이번주 기준의 SP에서 제대로 차감하지 않은 것이므로 빼줘야 함

    sp3 = sp1.loc[:, get_weekname_from(thisweek, 1):]
    sp4 = pd.concat([sr2, sp2, sp3], axis=1).fillna(0)
    sp4 = sp4.reset_index()
    sp4.insert(1, 'Country', sp4['To Site'].replace(site_map))
    sp4.insert(1, 'Region', sp4['Country'].replace(country_map))
    sp4.insert(0, 'Series', sp4['Mapping Model.Suffix'].apply(lambda x:x.split('-')[0]).replace(srt_model))
    sp4 = sp4.convert_dtypes()
    return sp4

def get_previous_weeklist(thisweek):
    lastweek = get_weekname_from(thisweek, -1)
    lastmonth = get_lastmonth(thisweek)
    weeklist1 = get_weeklist_for_certain_month(lastmonth[0], lastmonth[1])
    weeklist2 = get_weeklist_for_certain_month(datetime.date.fromisoformat(thisweek[:10]).isocalendar()[0] , get_month_from_date(datetime.date.fromisoformat(thisweek[:10])))
    if weeklist2.index(thisweek) == 0:
        previous_weeklist = weeklist1
    else:
        previous_weeklist = weeklist1 + weeklist2[:weeklist2.index(thisweek)]
    return previous_weeklist

def correct_gscp_data(today, frozen, vendor_list, ver):
    thisweek = get_weekname(today)
    frozen_period = [get_weekname_from(thisweek, i) for i in range(frozen+1)] # 확정 구간은 이번주를 포함하여 현재까지 확정되어 있는 주차들을 나타냄. 예를 들어 확정 구간 3주이면 당주 포함 총 4개 주차임
    previous_weeklist = get_previous_weeklist(thisweek)
    print(f'** {thisweek} 기준의 GSCP SP {ver} 로 작업합니다. 확정구간 : {frozen}주 **')
    sp_all = pd.read_excel(get_filename(), sheet_name=0, keep_default_na=False, na_values=['', 'nan']) # GSCP로부터 금주 SP를 데이타프레임화trim_column_name(sp)
    trim_column_name(sp_all)

    sp_total = pd.DataFrame()
    abnormal_sp = pd.DataFrame()

    for num, vendor in enumerate(vendor_list):
        print(f'-- 전체 업체 {len(vendor_list)} 개 중 {num+1} 번째, {vendor} 의 데이타 작업 시작합니다.--')
        print(f'    1. {vendor}의 GSCP raw 데이타 {ver} SP를 정리합니다.')
        c1 = sp_all['From Site'].str.contains(vendor.upper())
        c2 = sp_all['Category'].str.contains(ver)
        sp = sp_all[c1 & c2]
        sp = sp[['Mapping Model.Suffix', 'To Site']+get_pattern_from_list(sp.columns, re.compile('-\d\d-'))].reset_index(drop=True)
        sp = sp.convert_dtypes() # 적절한 데이타타입으로 변경
        sp['To Site'] = sp['To Site'].replace(site_name_adjust_map)
        sp = sp.groupby(['Mapping Model.Suffix', 'To Site']).sum()
        index_1 = sp.index # 추후 for 문을 돌릴 때 사용하기 위해 설정함

        sp = sp.reset_index()
        sp.insert(2, 'Category', 'SP')
        cat_order = CategoricalDtype(['SP', 'SR', 'PO', 'Real_SP'], ordered=True) # 데이타프레임을 순서대로 표시하기 위해 category 데이타로 설정
        sp['Category'] = sp['Category'].astype(cat_order)
        sp = sp.set_index(['Mapping Model.Suffix', 'To Site', 'Category'])

        for i in sp.index:
            for item in ['SR', 'PO', 'Real_SP']:
                sp.loc[i[0], i[1], item ] = 0

        sp = sp.sort_index(level=0)

        # 2. 선적 결과 DB에서 이전 주차의 선적값을 SR에 업데이트함
        print(f'    2. {vendor}의 선적 결과 DB 에서 이전 주차의 선적값을 SR에 업데이트합니다.')
        with open(f'D:/Data/{vendor} shipment result DB.bin', 'rb') as f:
            sr = pickle.load(f)
        sr = sr[sr['Week Name'].isin(previous_weeklist)]
        sr = sr.groupby(['Mapping Model.Suffix', 'Ship To', 'Week Name'])['Ship'].sum()
        for x, y, z in sr.index:
            sp.loc[(x, y, 'SR'), z] = sr[(x,y,z)]

        # 3. 지난 주 기준의 Open PO 를 통해, 금주 SP의 확정구간 Open PO 값 업데이트
        # 3-1) 확정 구간 중 이번주차의 Open PO 값 업데이트
        print(f'    3. {vendor}의 지난 주 기준 Open PO 를 계산하여 이번주 SP의 확정 구간의 PO값에 업데이트 합니다.')
        opo = get_open_po_at_certainweek(vendor, get_weekname_from(thisweek, -1))
        c1 = (opo['RSD Week Name'] <= thisweek) # PO의 RSD week가 이번주보다 같거나 작은 경우, 즉, 이번주에 해당하는 PO수량만 얻기 위해 필터링 조건임
        opo_th = opo[c1].pivot_table('Open_Qty', index=['Mapping Model.Suffix', 'Ship To'], columns='RSD Week Name', aggfunc=sum).sum(axis=1)

        for x, y in opo_th.index:
            sp.loc[(x, y, 'PO'), thisweek] = opo_th.loc[x, y]

        #  확정 구간 중, 이번 주 이후 구간의 Open PO값 업데이트
        c1 = (opo['RSD Week Name'].isin(frozen_period[1:]))
        opo_fu = opo[c1].pivot_table('Open_Qty', index=['Mapping Model.Suffix', 'Ship To'], columns='RSD Week Name', aggfunc=sum).fillna(0)

        for x, y in opo_fu.index:
            for week in opo_fu.columns:
                sp.loc[(x, y, 'PO'), week] = opo_fu.loc[(x, y), week]

        index_col = ['Type', 'Mapping Model.Suffix', 'To Site', 'Category']
        df_t = pd.DataFrame(columns=index_col)
        df_t = df_t.set_index(index_col)

        print(f'    4. {vendor}의 조건에 따른 Real SP를 업데이트하고 엑셀 파일로 출력합니다.')
        for x, y in index_1:
            t1 = sp.loc[(x, y, 'SP'), previous_weeklist].sum() - sp.loc[(x, y, 'SR'), previous_weeklist].sum() # 과거 주차들의 SP값과 SR값의 차이를 구한 값
            t2 = sp.loc[(x, y, 'SP'), thisweek] # 이번주차의 SP 값
            t3 = sp.loc[(x, y, 'PO'), thisweek] # 이번주차의 PO 값
            if (t1 == 0): # 과거 주차들의 SP 값과 SR 값의 차이가 없을 때는 실제 SP는 SR값이 되도록 함
                sp.loc[(x, y, 'Real_SP'), previous_weeklist] = sp.loc[(x, y, 'SR'), previous_weeklist]
                sum_sp = sp.loc[(x, y, 'SP'), frozen_period].sum() # 확정구간의 SP의 합계
                sum_po = sp.loc[(x, y, 'PO'), frozen_period].sum() # 확정구간의 PO의 합계
                if sum_sp == sum_po: # 확정 구간 SP와 PO 의 합이 같다면,
                    sp.loc[(x, y, 'Real_SP'), thisweek:] = sp.loc[(x, y, 'SP'), thisweek:] # 실제 SP 는 SP 값으로 반영함
                    reason = 'Normal Case'
                    for item in ['SP', 'SR', 'PO', 'Real_SP']:
                        t4 = sp.loc[(x, y, item)]
                        t4.name = (reason, x, y, item)
                        df_t = df_t.append(t4)
                else: # 확정 구간 SP와 PO 의 값이 다른 경우라면,
                    sp.loc[(x, y, 'Real_SP'), thisweek:] = sp.loc[(x, y, 'SP'), thisweek:]
                    reason = 'SP-PO is not matched in frozen period'
                    for item in ['SP', 'SR', 'PO', 'Real_SP']:
                        t4 = sp.loc[(x, y, item)]
                        t4.name = (reason, x, y, item)
                        df_t = df_t.append(t4)
            elif (t1 > 0): # 과거 주차들의 SP값이 SR 값보다 클 경우에는, 
                reason = 'SP is greater than SR'
                sp.loc[(x, y, 'Real_SP')] = sp.loc[(x, y, 'SP')]
                for item in ['SP', 'SR', 'PO', 'Real_SP']:
                    t4 = sp.loc[(x, y, item)]
                    t4.name = (reason, x, y, item)
                    df_t = df_t.append(t4)  
            elif (t1 < 0) & ((t1 + t2 - t3) == 0): # 과거 주차들의 SR 값이 SP 값보다 크고, 그 큰 폭이 이번주차의 SP값에서 이번주차 SP에서 이번주차 PO값을 차감한 값과 동일하다면, 이런경우는 실제 선적했는데 GSCP는 선적을 못한 것으로 carry over한 경우임
                sp.loc[(x, y, 'Real_SP'), previous_weeklist] = sp.loc[(x, y, 'SR'), previous_weeklist] # 과거주차의 실제 SP값은 SR 값으로 함
                sp.loc[(x, y, 'Real_SP'), thisweek] = t3 # 이번주차의 실제 SP값은 이번주차의 PO 값으로 함
                sum_sp = sp.loc[(x, y, 'SP'), frozen_period[1:]].sum()
                sum_po = sp.loc[(x, y, 'PO'), frozen_period[1:]].sum()
                if sum_sp == sum_po:
                    sp.loc[(x, y, 'Real_SP'), frozen_period[1]:] = sp.loc[(x, y, 'SP'), frozen_period[1]:]
#                     sp.loc[(x, y, 'Real_SP'), frozen_period[1:]] = sp.loc[(x, y, 'PO'), frozen_period[1:]]
                else:
                    sp.loc[(x, y, 'Real_SP'), frozen_period[1]:] = sp.loc[(x, y, 'SP'), frozen_period[1]:]
                    reason = 'SD input delay and SP-PO is not matched in frozen period'
                    for item in ['SP', 'SR', 'PO', 'Real_SP']:
                        t4 = sp.loc[(x, y, item)]
                        t4.name = (reason, x, y, item)
                        df_t = df_t.append(t4)
                    continue
                reason = 'SD input delay'
                for item in ['SP', 'SR', 'PO', 'Real_SP']:
                    t4 = sp.loc[(x, y, item)]
                    t4.name = (reason, x, y, item)
                    df_t = df_t.append(t4)
            elif (t1 < 0) & ((t1 + t2 - t3) != 0): # 과거 주차 SR 값이 SP 보다 큰데, 이번주차 SP에서 이번주차 PO값을 차감한 값과 다르다면, SP-PO의 동기화에 뭔가 문제가 있는 것으로 간주함
                reason = 'Something wrong on SP-PO'
#                 sum_sp = sp.loc[(x, y, 'SP'), previous_weeklist+frozen_period].sum()
#                 sum_sr_po = sp.loc[(x, y, 'SR'), previous_weeklist].sum() + sp.loc[(x, y, 'PO'), frozen_period].sum()
                sp.loc[(x, y, 'Real_SP'), previous_weeklist] = sp.loc[(x, y, 'SR'), previous_weeklist]
                sp.loc[(x, y, 'Real_SP'), thisweek:] = sp.loc[(x, y, 'SP'), thisweek:]
                for item in ['SP', 'SR', 'PO', 'Real_SP']:
                    t4 = sp.loc[(x, y, item)]
                    t4.name = (reason, x, y, item)
                    df_t = df_t.append(t4)

        sp = sp.reset_index()
        sp.insert(0, 'From Site', vendor)
        sp_total = pd.concat([sp_total, sp])
        if df_t.shape[0] != 0:
            df_t = df_t.reset_index()
            df_t.insert(0, 'From Site', vendor)
            abnormal_sp = pd.concat([abnormal_sp, df_t])
        #     df_t = df_t.append(['이상한 데이타가 없어요'])

    week_num = today.isocalendar()[1] # 이번주의 주차숫자(Week Number)를 구함
    update_time = datetime.datetime.now().strftime('%y%m%d%H')
    vendor_names = ''
    for name in vendor_list:
        vendor_names += '_' + name
    
    sp_total = sp_total.set_index(['From Site', 'Mapping Model.Suffix', 'To Site', 'Category'])
    abnormal_sp = abnormal_sp.set_index(['From Site', 'Type', 'Mapping Model.Suffix', 'To Site', 'Category'])
    filename = f"W{week_num}_GSCP_SP_correction_{vendor_names}_{update_time}_{ver}.xlsx" #이번주 forecast 파일을 저장할 이름 설정
    with pd.ExcelWriter("D:/Shipment Plan/GSCP Data/"+filename) as writer:
        sp_total.to_excel(writer, sheet_name='SP Analysis', merge_cells=False)
        abnormal_sp.to_excel(writer, sheet_name='abnormal SP list', merge_cells=False)

def get_sp_from_GSCP_DB(thisweek, ver, vendor, updated_ver=-1):
    with open('D:/Data/GSCP raw data.bin', 'rb') as f:
        gscp = pickle.load(f)

    c1 = (gscp['Ref'] == thisweek)
    c2 = (gscp['Ver'] == ver)
    c3 = (gscp['From Site'] == vendor)
    updated_time = gscp[c1 & c2 & c3]['Updated_at'].unique()[updated_ver]
    c4 = (gscp['Updated_at'] == updated_time)
    sp_thisweek = gscp[c1 & c2 & c3 & c4]
    sp_thisweek = sp_thisweek.groupby(['Mapping Model.Suffix', 'To Site']).sum(numeric_only=True).drop(columns='Frozen').sort_index(axis=1)
    sp_thisweek = sp_thisweek.loc[:, get_previous_weeklist(thisweek)[0]:]

    sp_thisweek['Sum'] = sp_thisweek.sum(axis=1)
    sp_thisweek = sp_thisweek[sp_thisweek['Sum'] > 0]
    sp_thisweek = sp_thisweek.drop(columns='Sum')

    sp_thisweek = sp_thisweek.reset_index()
    sp_thisweek.insert(1, 'Country', sp_thisweek['To Site'].replace(site_map))
    sp_thisweek.insert(1, 'Region', sp_thisweek['Country'].replace(country_map))
    if vendor in ['Quanta', 'Pegatron', 'Wingtech']:
        sp_thisweek.insert(0, 'Series', sp_thisweek['Mapping Model.Suffix'].apply(lambda x:x.split('-')[0]).replace(srt_model))
    return sp_thisweek

def make_forecast_change_for_certain_model(ref_point, period, model, ver, vendor):
    df_rst = pd.DataFrame()
    ref_week = get_weekname_from(get_weekname(datetime.date.today()), ref_point)
    for n in range(0, -period, -1):
        week = get_weekname_from(ref_week, n)
        df = category_sum(get_sp_from_GSCP_DB(week, ver, vendor).groupby('Series').sum(numeric_only=True))
        df.loc['Total'] = df.sum()
        sr = df.loc[model]
        df_rst[week] = sr
    df_rst = df_rst.sort_index().fillna(0)
    return df_rst

def plot_forecast_change(df, name):
    plt.figure(figsize=(15,8))
    plt.title(name, size=20)
    plt.plot(df.index, df.iloc[:, 3].tolist(), color='gray', ls='--', lw=1, marker='x', alpha=0.6, label=df.columns[3])
    plt.plot(df.index, df.iloc[:, 2].tolist(), color='darkgray', lw=4, marker='v', alpha=0.6, label=df.columns[2])
    plt.plot(df.index, df.iloc[:, 1].tolist(), color='dimgray', lw=6, marker='s', alpha=0.9, label=df.columns[1])
    plt.plot(df.index, df.iloc[:, 0].tolist(), color='k', lw=8, marker='o', markersize=13, label=df.columns[0])
    plt.legend()
    plt.ylabel('QTY', rotation=0, loc='top')
    plt.xlabel('Month')
    plt.grid(axis='y')
    plt.savefig('D:/figure/forecast_change_during_4weeks.png')

def get_DPK_stock():
    # DPK 재고의 가장 최신 재고의 산정 기준날짜 확인
    with open('D:/Data/DPK stock.bin', 'rb') as f:
        dpk_stock = pickle.load(f)

    # 재고 산정 기준 일자의 월부터 오늘날짜의 월까지의 Quanta의 월별 생산계획을 취합하기 위해 계산해야 할 월의 명단을 리스트로 만듬
    stock_ref_date = dpk_stock['Ref_date'].max().date()
    months = get_months_between_dates(stock_ref_date, datetime.date.today())
    dpk_stock = dpk_stock[dpk_stock['Ref_date'] == stock_ref_date.strftime('%Y-%m-%d')]

    # 취합할 생산계획 데이타프레임의 초기값을 빈 프레임으로 만들어 놓고 해당 월의 데이타를 누적시켜 생산계획을 취합함
    pr = pd.DataFrame()
    for target_month in months:
        if datetime.datetime.strftime(target_month, '%Y-%m') == datetime.datetime.strftime(stock_ref_date, '%Y-%m'):
            inp_plan = get_quanta_input_qty_by_month(target_month.year, target_month.month)
            inp_plan = inp_plan[inp_plan['Input Date'].dt.month == target_month.month]
            inp_plan = inp_plan[inp_plan['Input Date'] >= datetime.datetime.strftime(stock_ref_date, '%Y-%m-%d')]
            pr = pd.concat([pr, inp_plan]).reset_index(drop=True)
        else:
            inp_plan = get_quanta_input_qty_by_month(target_month.year, target_month.month)
            inp_plan = inp_plan[inp_plan['Input Date'].dt.month == target_month.month]
            pr = pd.concat([pr, inp_plan]).reset_index(drop=True)

    i_num = [get_weekname_from(pr['LG Week'].min(), n) for n in range(0, 30)].index(pr['LG Week'].max())
    cols = [get_weekname_from(pr['LG Week'].min(), n) for n in range(0, 30)][:i_num+1]
    pr_by_os = pd.DataFrame(0, index=pr['Mapping Model.Suffix'].unique(), columns=cols)
    pr = pr.pivot_table('QTY', index='Mapping Model.Suffix', columns='LG Week', aggfunc=sum).fillna(0)
    for model in pr.index:
        for week in pr.columns:
            pr_by_os.loc[model, week] = pr.loc[model, week]
    pr_by_os.index.name = 'Mapping Model.Suffix'
    # 취합된 생산계획 데이타 프레임을 LG 주차별 생산계획 테이블로 전환하여  DPR의 OS 정보를 가져와 병합하여 OS별 생산계획으로 표현하고 여기에 DPK 기초 재고 정보도 추가함
    col = ['Model.Suffix', 'OS TYPE']
    pr_by_os = pr_by_os.reset_index().merge(get_pdr()[col], left_on='Mapping Model.Suffix', right_on='Model.Suffix').drop(columns='Model.Suffix').groupby('OS TYPE').sum(numeric_only=True).reset_index()

    dpk_stock = pd.merge(pr_by_os, dpk_stock, how='outer').fillna(0).drop(columns='Ref_date')
    dpk_stock['MS P/N']= dpk_stock['OS TYPE'].map(dpk_map2)
    
    # DPK PO 정보를 만들어 둠
    with open('D:/Data/DPK blanket PO DB.bin', 'rb') as f:
            bpo = pickle.load(f)
    bpo = bpo.loc[bpo['LG PO Date'] >= datetime.datetime.strftime(stock_ref_date, '%Y-%m-%d'), ('LG PO Date', 'MS P/N', 'Order Qty')].groupby(['LG PO Date', 'MS P/N']).sum()
    bpo = bpo.reset_index()
    bpo['LG Week'] = bpo['LG PO Date'].apply(get_weekname)
    bpo_pv = bpo.pivot_table('Order Qty', index='MS P/N', columns='LG Week')

    col1 = ['OS'] * 2 + get_weeklist(dpk_stock) * 3
    col2 = ['NAME', 'P/N'] + ['BOH'] * len(get_weeklist(dpk_stock)) + ['Input'] * len(get_weeklist(dpk_stock)) + ['DPK PO'] * len(get_weeklist(dpk_stock)) 
    
    # 각 주차별 생산/DPK 주문정보를 바탕으로 DPK 재고 정보를 계산함
    dpk_stock_balance = pd.DataFrame(data=0, index=range(len(dpk_stock['OS TYPE'])), columns=[col1, col2])
    dpk_stock_balance.loc[:, ('OS', 'NAME')] = dpk_stock['OS TYPE']
    dpk_stock_balance.loc[:, ('OS', 'P/N')] = dpk_stock['MS P/N']
    dpk_stock_balance.loc[:, (get_weeklist(dpk_stock)[0], 'BOH')] = dpk_stock['Stock']
    for week in get_weeklist(dpk_stock):
        dpk_stock_balance.loc[:, (week, 'Input')] = dpk_stock[week]
    for pn in bpo_pv.index:
        for week in bpo_pv.columns:
            c1 = (dpk_stock_balance[('OS', 'P/N')] == pn)
            dpk_stock_balance.loc[c1, (week, 'DPK PO')] = bpo_pv.loc[pn, week]
    dpk_stock_balance = dpk_stock_balance.fillna(0).convert_dtypes()

    for week in get_weeklist(dpk_stock)[1:]:
        for dpk in dpk_stock['MS P/N'].dropna():
            c1 = (dpk_stock_balance[('OS', 'P/N')] == dpk)
            dpk_stock_balance.loc[c1, (week, 'BOH')] = dpk_stock_balance.loc[c1, (get_weekname_from(week, -1) , 'BOH')].iloc[0] - dpk_stock_balance.loc[c1, (get_weekname_from(week, -1), 'Input')].iloc[0] + dpk_stock_balance.loc[c1, (get_weekname_from(week, -1), 'DPK PO')].iloc[0]
    return dpk_stock_balance.set_index([('OS', 'NAME'), ('OS', 'P/N')]).sort_index(axis=1)

def get_quanta_boh(date): # 특정 날짜에 해당하는 월의 첫번째 주차 기준의 boh를 가져오는 함수
    with open('D:/Data/boh_db.bin', 'rb') as f:
        boh_quanta = pickle.load(f)
    return boh_quanta[boh_quanta['Ref'] == get_weeklist_for_certain_month(date.year, get_month_from_date(date))[0]].reset_index(drop=True)

def move_month(year, month, jump):
    year = year + ((month-1) + jump)//12
    month = (month + jump) % 12
    if month != 0:
        month
    else:
        month = 12
    return year, month

def get_shipment_result(vendor, num):
    with open(f'D:/Data/{vendor} shipment result DB.bin', 'rb') as f:
        df = pickle.load(f)
    week_name = get_weekname_from(get_weekname(datetime.date.today()), num)
    return df[df['Week Name'] == week_name].pivot_table('Ship', index=['Country', 'BL No', 'BL Status', 'Series', 'Mapping Model.Suffix'], columns='Ship Date', aggfunc=sum).fillna(0).convert_dtypes()

def get_months_between_dates(date1, date2):
    """
    date1과 date2 사이의 모든 월의 정보를 출력하는 함수
    :param date1: datetime 객체. 시작 날짜
    :param date2: datetime 객체. 종료 날짜
    :return: datetime 객체 리스트. date1과 date2 사이의 모든 월
    """

    # date1과 date2 중에서 더 이른 날짜를 시작 날짜로 설정
    if date1 > date2:
        start_date = date2
        end_date = date1
    else:
        start_date = date1
        end_date = date2

    # 시작 날짜의 월부터 종료 날짜의 월까지 datetime 객체 리스트 생성
    current_month = start_date.replace(day=1)
    end_month = end_date.replace(day=1)
    months_between = []
    while current_month <= end_month:
        months_between.append(current_month)
        current_month += relativedelta(months=1)

    return months_between

def get_quanta_input_qty_by_month(year, month):
    with open('D:/Data/Quanta Input Plan.bin', 'rb') as f:
        pr = pickle.load(f)
    c1 = pd.to_datetime(pr['Created_at']).dt.year == year
    c2 = pd.to_datetime(pr['Created_at']).dt.month == month
    input_day = pr[c1 & c2]['Created_at'].max()
    pr = pr[pr['Created_at'] == input_day]
    return pr[pr['Input Date'].dt.month == month].reset_index(drop=True)

def get_sp_po_gap(vendor_list, confirm_period=5):
    dir_path = 'D:/Shipment Plan/GSCP Data/'
    list_of_files = [file for file in os.listdir(dir_path) if 'GSCP_SP_correction' in file]
    fn = max([os.path.join(dir_path, filename) for filename in list_of_files], key=os.path.getctime)
    df = pd.read_excel(fn, sheet_name='SP Analysis')
    df = df[df['From Site'].isin(vendor_list)]
    df['Series'] = df['Mapping Model.Suffix'].apply(lambda x:x.split('-')[0]).replace(srt_model)
    confirm_weeks = [get_weekname_from(get_weekname(datetime.date.today()), i) for i in range(confirm_period)]
    df = df[df['Category'].isin(['Real_SP', 'PO'])].groupby(['Series', 'Mapping Model.Suffix', 'To Site', 'Category']).sum(numeric_only=True)[confirm_weeks].unstack()
    df = df[~(df.sum(axis=1)==0)]
    df[('SUM', 'PO')] = df.xs('PO', axis=1, level=1).sum(axis=1)
    df[('SUM', 'Real_SP')] = df.xs('Real_SP', axis=1, level=1).sum(axis=1)
    df[('SUM', 'GAP')] = df[('SUM', 'PO')] - df[('SUM', 'Real_SP')]
    return df

def get_boh_nextweek(boh_thisweek, thisweek):
    search_month = str(datetime.date.fromisoformat(thisweek[:10]).isocalendar().year) + str(get_month_from_date(datetime.date.fromisoformat(thisweek[:10]))).zfill(2)

    with open('D:/Data/Quanta Input Plan.bin', 'rb') as f:
        input_df = pickle.load(f)

    final_input_date = input_df[input_df['Created_at'].apply(lambda x:x.strftime('%Y%m')).isin([search_month])]['Created_at'].max()
    input_df = input_df[(input_df['Created_at']==final_input_date) & (input_df['LG Week'] == thisweek)]

    with open('D:/Data/Quanta shipment result DB.bin', 'rb') as f:
        sr = pickle.load(f)
    sr = sr[sr['Week Name']==thisweek]

    boh_nextweek = boh_thisweek.merge(input_df.groupby('Mapping Model.Suffix')['QTY'].sum().reset_index(), how='outer').merge(sr.groupby('Mapping Model.Suffix')['Ship'].sum().reset_index(), how='outer').fillna(0)
    boh_nextweek['BOH'] = boh_nextweek['BOH'] + boh_nextweek['QTY'] - boh_nextweek['Ship']
    return boh_nextweek[boh_nextweek['BOH'] > 0].drop(['QTY', 'Ship'], axis=1).reset_index(drop=True)

def update_glop_shipment():
    filepath = 'D:/Downloads/' + [file for file in os.listdir('D:/Downloads/') if 'poshThru' in file][-1] # 파일이름에 poshThru가 들어가는 파일 중에 가장 최신 파일을 선택함
    df = pd.read_excel(filepath, sheet_name='DataSet').iloc[:, :48]

    if df['Model'][0].split('.')[0] == 'PH30N':
        vendor = 'Wanlida'
    else:
        vendor = df['Model'].apply(lambda x:x.split('-')[0]).replace(srt_model).replace(vendor_find)[0]

    if vendor not in ['Quanta', 'Pegatron', 'Wanlida', 'Wingtech']:
        sys.exit()

    df = df[['Model', 'PO No.', 'Ship To', 'Shipping', 'Cancel', 'PO', 'OQC Report', 'OQC Date', 'OQC Result', 'Ship', 'Issued Date', 'RSD','Ship Date', 'BL No', 'BL Status', 'Method', 'Price Term', 'Unit Price', 'Currency', 'Payment Term', 'PO1 No', 'SO No.', 'Final Destination']]
    df.loc[df['PO1 No'].notnull(), 'PO1 No'] = df.loc[df['PO1 No'].notnull(),'PO1 No'].astype('int').astype('str')
    df.loc[df['SO No.'].notnull(), 'SO No.'] = df.loc[df['SO No.'].notnull(), 'SO No.'].astype('int').astype('str')
    df[['OQC Date', 'PO1 No', 'SO No.', 'Final Destination']] = df[['OQC Date', 'PO1 No', 'SO No.', 'Final Destination']].fillna('-')

    df['Ship'].fillna(0, inplace=True)
    df['Ship'] = df['Ship'].astype('int')
    df['Issued Date'] = pd.to_datetime(df['Issued Date'])
    df['PO Week'] = df['Issued Date'].dt.isocalendar().week.apply(lambda x:'{0:02d}'.format(x))
    df['PO Month'] = df['Issued Date'].apply(get_month_from_date)
    df['PO Year'] = df['Issued Date'].dt.isocalendar().year
    df['RSD'] = pd.to_datetime(df['RSD']) # PO의 RSD를 Datetime 형식으로 변경
    df['RSD Week'] = df['RSD'].dt.isocalendar().week.apply(lambda x:'{0:02d}'.format(x))
    df['RSD Month'] = df['RSD'].apply(get_month_from_date)
    df['RSD Year'] = df['RSD'].dt.isocalendar().year
    df['RSD Week Year'] = df['RSD'].apply(getfirstdate_year)
    df['RSD Week Month'] = df['RSD'].apply(getfirstdate_month)
    df['RSD Week Day'] = df['RSD'].apply(getfirstdate_day)
    df['RSD Week Name'] = df['RSD Week Year'].astype(str) + '-' + df['RSD Week Month'].astype(str) + '-' + df['RSD Week Day'].astype(str) + '(W' + df['RSD Week'].astype(str) +')'
    df.rename(columns={'Model':'Mapping Model.Suffix'}, inplace=True)

    if vendor != 'Wanlida':
        df['Model'] = df['Mapping Model.Suffix'].apply(lambda x: x.split('-')[0])
        df['Series'] = df['Model'].replace(srt_model)

    df['Ship To'] = df['Ship To'].replace(site_name_adjust_map)
    df['Country'] = df['Ship To'].replace(site_map)
    df['Region'] = df['Country'].replace(country_map)

    if vendor != 'Wanlida':
        df1 = df.groupby(['PO No.', 'Cancel', 'Issued Date', 'PO Year', 'PO Month', 'PO Week', 'Region', 'Country', 'Series', 'Model',
                            'Mapping Model.Suffix', 'Ship To', 'RSD', 'RSD Year', 'RSD Month', 'RSD Week', 'RSD Week Name', 'Method',
                            'Price Term', 'Unit Price', 'Currency', 'Payment Term', 'PO1 No', 'SO No.',
                            'Final Destination'])[['PO']].sum().reset_index().set_index('PO No.')
    else:
        df1 = df.groupby(['PO No.', 'Cancel', 'Issued Date', 'PO Year', 'PO Month', 'PO Week', 'Region', 'Country',
                        'Mapping Model.Suffix', 'Ship To', 'RSD', 'RSD Year', 'RSD Month', 'RSD Week', 'RSD Week Name', 'Method',
                        'Price Term', 'Unit Price', 'Currency', 'Payment Term', 'PO1 No', 'SO No.',
                        'Final Destination'])[['PO']].sum().reset_index().set_index('PO No.')  

    with open(f'D:/Data/{vendor} PO DB.bin', 'rb') as f:
        PO_df = pickle.load(f)
    PO_df1 = PO_df.copy()
    prev_po = PO_df.shape[0]
    for i in df1.index:
        PO_df.loc[i, :] = df1.loc[i, :]
    updated_po = PO_df.shape[0]
    with open(f'D:/Data/{vendor} PO DB.bin', 'wb') as f:
        pickle.dump(PO_df, f)
    # DB backup 폴더에도 저장
    with open(f'D:/Data/DB backup/{vendor} PO DB.bin', 'wb') as f:
        pickle.dump(PO_df, f)

    with open(f'D:/Data/{vendor} shipment result DB.bin', 'rb') as f:
        SR_df = pickle.load(f)

    df = df[df['Ship Date'].notnull()]
    df['Ship Date'] = pd.to_datetime(df['Ship Date'])
    df['Ship Week'] = df['Ship Date'].dt.isocalendar().week.apply(lambda x:'{0:02d}'.format(x))
    df['Ship Month'] = df['Ship Date'].apply(get_month_from_date)
    df['Ship Year'] = df['Ship Date'].dt.isocalendar().year
    df['Ship Week month'] = df['Ship Date'].apply(getfirstdate_month)
    df['Ship Week day'] = df['Ship Date'].apply(getfirstdate_day)
    df['Ship Week year'] = df['Ship Date'].apply(getfirstdate_year)
    df['Ship Week year'] = df['Ship Week year'].astype(str)
    df['Ship Week month'] = df['Ship Week month'].astype(str)
    df['Ship Week day'] = df['Ship Week day'].astype(str)
    df['Ship Week'] = df['Ship Week'].astype(str)
    df['Week Name'] = df['Ship Week year'] + '-' + df['Ship Week month'] + '-' + df['Ship Week day'] + '(W' + df['Ship Week'] + ')'
    df = df.set_index(['PO No.', 'BL No'])
    SR_df1 = SR_df.copy()
    prev_sr = SR_df.shape[0]
    for i in df.index:
        SR_df.loc[i, :] = df.loc[i, :]
    updated_sr = SR_df.shape[0]

    # PO 대비 선적수량이 더 많은 경우가 없는지 check
    with open(f'D:/Data/{vendor} PO DB.bin', 'rb') as f:
        PO = pickle.load(f)
    with open(f'D:/Data/{vendor} shipment result DB.bin', 'rb') as f:
        SR = pickle.load(f)

    PO = PO[PO['Cancel']=='N']
    df = pd.concat([PO['PO'], SR.groupby('PO No.')['Ship'].sum()], axis=1)
    df.fillna(0, inplace=True)
    df['Open'] = df['PO'] - df['Ship']
    df = df[df['Open'] < 0]
    if df.shape[0] != 0:
        print(df)
        sys.exit()
    with open(f'D:/Data/{vendor} shipment result DB.bin', 'wb') as f:
        pickle.dump(SR_df, f)
    with open(f'D:/Data/DB backup/{vendor} shipment result DB.bin', 'wb') as f:
        pickle.dump(SR_df, f)
    print(f"{vendor}의 DB를 업데이트 하였습니다. PO는 {updated_po - prev_po} 건, 선적은 {updated_sr - prev_sr} 건 추가 되었습니다. ")
    print(filepath)

    print('추가된 PO 현황입니다.')
    print(PO_df.drop(index=PO_df1.index).pivot_table('PO', index=['PO No.', 'Mapping Model.Suffix',  'Country'], columns='RSD Week Name', aggfunc=sum)) # 추가된 PO 리스트
    print('추가된 선적 현황입니다.')
    print(SR_df.drop(index=SR_df1.index).pivot_table('Ship', index=['Country', 'BL No', 'Mapping Model.Suffix', 'PO No.'], columns=['Week Name', 'Ship Date'], aggfunc='sum').fillna(0)) # 추가된 선적실적

def download_excel_on_GLOPP_at_PT(vendor):
    pyautogui.click(444, 23) # 주소 표시줄 클릭
    pyautogui.sleep(0.3)
    pyautogui.typewrite('glopp.lge.com/index.jsp', interval=0.1)
    pyautogui.press('enter', presses=2, interval=1)
    pyautogui.sleep(1)
    while not pyautogui.pixelMatchesColor(868, 619, (255, 255, 255)):  # GLOPP 최초 로그인 후 초기 정보 로딩까지 대기
        pyautogui.sleep(1)
    pyautogui.moveTo(926, 592, duration=1)
    pyautogui.click(1613, 121, duration=0.2) # 업체 선택 dropdown 메뉴 클릭
    pyautogui.moveTo(1715, 160, duration=0.5) # scroll bar를 클릭

    if vendor == 'Pegatron PC':
        pyautogui.dragRel(0, 30, duration=0.3) # scroll bar를 내림
        pyautogui.click(1550, 179, duration=0.2) # Pegatron PC 선택
    elif vendor == 'Pegatron Thin Client':
        pyautogui.dragRel(0, 30, duration=0.3) # scroll bar를 내림
        pyautogui.click(1551, 198, duration=0.2) # Pegatron Thin Client 선택
    elif vendor == 'Quanta':
        pyautogui.dragRel(0, 40, duration=0.3) # scroll bar를 내림
        pyautogui.click(1511, 189, duration=0.2) # Quanta 선택
    elif vendor == 'Wingtech':
        pyautogui.dragRel(0, 60, duration=0.3) # scroll bar를 내림
        pyautogui.click(1530, 223, duration=0.2) # Wingtech 선택
    elif vendor == 'TCL MOKA':
        Point(x=1539, y=196)
        pyautogui.dragRel(0, 50, duration=0.3) # scroll bar를 내림
        pyautogui.click(1550, 159, duration=0.2) # TCL MOKA 선택

    pyautogui.sleep(1)
    while not pyautogui.pixelMatchesColor(899, 905, (255, 255, 255)):  # 업체 선택 후 조회될때까지
        pyautogui.sleep(1)
    pyautogui.moveTo(613, 82, duration=0.5) # Shipping & Invoicing 선택
    pyautogui.click(610, 121, duration=0.5) # Shipping & Invoicing 하위 메뉴 선택
    pyautogui.sleep(1)
    while not pyautogui.pixelMatchesColor(276, 1407, (255, 255, 255)):  # Shipping & Invoicing 메뉴로 전환될 때까지
        pyautogui.sleep(1)
    pyautogui.click(1884, 186, duration=0.1)
    pyautogui.press('enter')
    print('GLOP에서 파일 다운로드를 시작합니다.')
    while not pyautogui.pixelMatchesColor(1350, 58, (74, 74, 74)):  # 파일 다운로드 완료될 때까지
        pyautogui.sleep(1)

def get_tcl_boh(target_week):
    with open('D:/Data/TCL PP_result.bin', 'rb') as f:
        ppr = pickle.load(f)
    models = ppr['Mapping Model.Suffix'].unique()
    
    with open('D:/Data/ODM_SR.db', 'rb') as f:
        sr = pickle.load(f)
        sr = sr.loc['OS_TCL_CN_P']

    sr['Week Name'] = sr['Ship Date'].map(get_weekname)
    sr = sr[['Mapping Model.Suffix', 'Ship', 'Week Name']].reset_index(drop=True)
    sr = sr.groupby(['Mapping Model.Suffix','Week Name']).sum('Ship').reset_index()
    c1 = sr['Mapping Model.Suffix'].isin(models)
    c2 = sr['Week Name'] < target_week

    ppr['Week Name'] = ppr['PP Date'].map(get_weekname)
    ppr = ppr.groupby(['Mapping Model.Suffix', 'Rep PMS', 'TCL BOM', 'Week Name']).sum('QTY').reset_index()
    c3 = ppr['Mapping Model.Suffix'].isin(models)
    c4 = ppr['Week Name'] < target_week
    boh = ppr[c3 & c4].groupby(['Mapping Model.Suffix', 'Rep PMS', 'TCL BOM']).sum('QTY').reset_index().merge(sr[c1 & c2].groupby('Mapping Model.Suffix').sum('Ship').reset_index(), how='outer').fillna(0)
    boh = boh.rename(columns={'QTY':'PP'})
    boh['BOH'] = boh['PP'] - boh['Ship']
    return boh